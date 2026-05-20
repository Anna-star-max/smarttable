import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import uuid

class ExcelGenerator:
    def __init__(self, output_dir: str = "./output"):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True, name="微軟正黑體", size=11)
        self.label_font = Font(bold=True, name="微軟正黑體", size=11)
        self.content_font = Font(name="微軟正黑體", size=11)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def auto_size_column(self, ws, min_width=12, max_width=60):
        for col in range(1, ws.max_column + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            for row in range(1, ws.max_row + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            adjusted_width = min(max(max_length + 4, min_width), max_width)
            ws.column_dimensions[column_letter].width = adjusted_width

    def auto_size_rows(self, ws):
        for row in range(1, ws.max_row + 1):
            max_height = 20
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    lines = str(cell.value).count('\n') + 1
                    height = lines * 15
                    if height > max_height:
                        max_height = height
            ws.row_dimensions[row].height = max_height

    def generate(self, data: dict, meeting_topic: str = "未命名會議") -> str:
        content_type = data.get("content_type", "meeting")
        data_table = data.get("data_table")
        bug_report = data.get("bug_report")
        if content_type == "data_table" or (data_table and data_table.get("headers") and data_table.get("rows")):
            return self.generate_data_table(data, meeting_topic)
        elif content_type == "bug_report" or (bug_report and bug_report.get("headers") and bug_report.get("rows")):
            return self.generate_bug_report(data, meeting_topic)
        else:
            return self.generate_meeting(data, meeting_topic)

    def generate_bug_report(self, data: dict, meeting_topic: str = "Bug問題分析") -> str:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.sheet_view.showGridLines = False

        bug_info = data.get("bug_report") or {}
        headers = bug_info.get("headers") or []
        rows = bug_info.get("rows") or []
        title = bug_info.get("標題") or meeting_topic
        summary = bug_info.get("摘要") or ""
        findings = bug_info.get("關鍵發現") or []
        suggestions = bug_info.get("建議") or ""

        # 第一行：標題
        ws.cell(row=1, column=1, value=title).font = Font(bold=True, name="微軟正黑體", size=14)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
        if headers:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

        # 第二行：摘要
        if summary:
            ws.cell(row=2, column=1, value="摘要").font = self.label_font
            ws.cell(row=2, column=1).border = self.border
            ws.cell(row=2, column=1).alignment = Alignment(horizontal="center", vertical="center")
            summary_cell = ws.cell(row=2, column=2, value=summary)
            summary_cell.font = self.content_font
            summary_cell.border = self.border
            summary_cell.alignment = Alignment(vertical="center", wrap_text=True)
            if headers:
                ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=len(headers))
            ws.row_dimensions[2].height = 30

        # 關鍵發現
        if findings:
            row = 3 if not summary else 4
            ws.cell(row=row, column=1, value="關鍵發現").font = self.label_font
            ws.cell(row=row, column=1).border = self.border
            ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")
            findings_cell = ws.cell(row=row, column=2, value="\n".join(findings))
            findings_cell.font = self.content_font
            findings_cell.border = self.border
            findings_cell.alignment = Alignment(vertical="top", wrap_text=True)
            if headers:
                ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=len(headers))
            ws.row_dimensions[row].height = max(len(findings) * 20, 25)

        # 建議
        if suggestions:
            row = 4 if (not summary and not findings) else (5 if not summary else row + 1)
            ws.cell(row=row, column=1, value="建議").font = self.label_font
            ws.cell(row=row, column=1).border = self.border
            ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")
            sugg_cell = ws.cell(row=row, column=2, value=suggestions)
            sugg_cell.font = self.content_font
            sugg_cell.border = self.border
            sugg_cell.alignment = Alignment(vertical="center", wrap_text=True)
            if headers:
                ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=len(headers))
            ws.row_dimensions[row].height = 30

        # 欄位標題行
        header_row = 6 if (summary or findings or suggestions) else 2
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=h)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self.border

        # 資料列
        for row_idx, row_data in enumerate(rows, header_row + 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = self.content_font
                cell.border = self.border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # 自動調整欄寬
        for col in range(1, len(headers) + 1):
            col_letter = get_column_letter(col)
            max_length = 0
            for row in range(header_row, ws.max_row + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    cell_length = len(str(cell.value)) * 1.2
                    if cell_length > max_length:
                        max_length = cell_length
            adjusted_width = max(12, min(max_length + 4, 60))
            ws.column_dimensions[col_letter].width = adjusted_width

        # 設置列高
        for row in range(1, ws.max_row + 1):
            ws.row_dimensions[row].height = 20

        # 生成檔名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Bug報告_{timestamp}_{uuid.uuid4().hex[:6]}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        wb.save(filepath)
        return filename

    def generate_meeting(self, data: dict, meeting_topic: str = "未命名會議") -> str:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.sheet_view.showGridLines = False

        overview = data.get("meeting_overview") or {}
        action_items = data.get("action_items") or []

        # 第一行：主題
        c1 = ws.cell(row=1, column=1, value="主題")
        c1.font = self.label_font
        c1.border = self.border
        c1.alignment = Alignment(horizontal="center", vertical="center")
        c2 = ws.cell(row=1, column=2, value=overview.get("主題") or "-")
        c2.font = self.content_font
        c2.border = self.border
        c2.alignment = Alignment(vertical="center", wrap_text=True)

        # 第二行：日期
        c3 = ws.cell(row=2, column=1, value="日期")
        c3.font = self.label_font
        c3.border = self.border
        c3.alignment = Alignment(horizontal="center", vertical="center")
        c4 = ws.cell(row=2, column=2, value=overview.get("日期") or "-")
        c4.font = self.content_font
        c4.border = self.border
        c4.alignment = Alignment(vertical="center")

        # 第三行：與會人員
        c5 = ws.cell(row=3, column=1, value="與會人員")
        c5.font = self.label_font
        c5.border = self.border
        c5.alignment = Alignment(horizontal="center", vertical="center")
        c6 = ws.cell(row=3, column=2, value=", ".join(overview.get("與會人員", []) or []) or "-")
        c6.font = self.content_font
        c6.border = self.border
        c6.alignment = Alignment(vertical="center")

        # 第四行：重點摘要
        c7 = ws.cell(row=4, column=1, value="重點摘要")
        c7.font = self.label_font
        c7.border = self.border
        c7.alignment = Alignment(horizontal="center", vertical="center")
        c8 = ws.cell(row=4, column=2, value=overview.get("重點摘要") or "-")
        c8.font = self.content_font
        c8.border = self.border
        c8.alignment = Alignment(vertical="center", wrap_text=True)

        # 第五行：空白列
        ws.row_dimensions[5].height = 20

        # 第六行：行動項目表頭
        row_start = 6
        action_headers = ["行動項目", "負責人", "狀態", "截止日期"]
        for col, header in enumerate(action_headers, 1):
            cell = ws.cell(row=row_start, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self.border

        # 行動項目內容
        if action_items:
            for row_idx, item in enumerate(action_items, row_start + 1):
                c_item = ws.cell(row=row_idx, column=1, value=item.get("項目名稱") or "-")
                c_item.font = self.content_font
                c_item.border = self.border
                c_item.alignment = Alignment(vertical="center", wrap_text=True)

                c_owner = ws.cell(row=row_idx, column=2, value=item.get("負責人") or "-")
                c_owner.font = self.content_font
                c_owner.border = self.border
                c_owner.alignment = Alignment(horizontal="center", vertical="center")

                c_status = ws.cell(row=row_idx, column=3, value=item.get("狀態", "pending"))
                c_status.font = self.content_font
                c_status.border = self.border
                c_status.alignment = Alignment(horizontal="center", vertical="center")

                c_due = ws.cell(row=row_idx, column=4, value=item.get("截止日期") or "-")
                c_due.font = self.content_font
                c_due.border = self.border
                c_due.alignment = Alignment(horizontal="center", vertical="center")
        else:
            c_none = ws.cell(row=row_start + 1, column=1, value="無行動項目")
            c_none.font = self.content_font
            c_none.border = self.border
            c_none.alignment = Alignment(horizontal="center", vertical="center")

        # 自動調整欄寬（根據內容）
        for col in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col)
            max_length = 0
            for row in range(1, ws.max_row + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    # 中文字元寬度計算
                    cell_length = len(str(cell.value)) * 1.2
                    if cell_length > max_length:
                        max_length = cell_length
            # 設置範圍 12-80
            adjusted_width = max(12, min(max_length + 4, 80))
            ws.column_dimensions[col_letter].width = adjusted_width

        # 設置列高（根據內容和欄寬計算換行行數）
        for row in range(1, ws.max_row + 1):
            max_lines = 1
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value and cell.alignment and cell.alignment.wrap_text:
                    # 計算預估換行行數：內容長度 / 欄寬（字元數）
                    col_letter = get_column_letter(col)
                    col_width = ws.column_dimensions[col_letter].width
                    if col_width > 0:
                        # 中文字約佔 1.5 單位寬度
                        chars_per_line = int(col_width / 1.5)
                        if chars_per_line > 0:
                            content = str(cell.value)
                            lines = max(1, (len(content) // chars_per_line) + 1)
                            if lines > max_lines:
                                max_lines = lines
            ws.row_dimensions[row].height = max(max_lines * 18, 20)

        # 生成檔名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"會議記錄_{timestamp}_{uuid.uuid4().hex[:6]}.xlsx"
        filepath = os.path.join(self.output_dir, filename)

        wb.save(filepath)
        return filename

    def generate_data_table(self, data: dict, title: str = "數據表格") -> str:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.sheet_view.showGridLines = False

        data_table = data.get("data_table") or {}
        headers = data_table.get("headers") or []
        rows = data_table.get("rows") or []
        table_title = data_table.get("title") or title

        # 第一行：表格標題
        ws.cell(row=1, column=1, value=table_title).font = Font(bold=True, name="微軟正黑體", size=14)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
        if headers:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

        # 第二行：欄位標題
        header_row = 2
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self.border

        # 資料列
        for row_idx, row_data in enumerate(rows, header_row + 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = self.content_font
                cell.border = self.border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # 總結區域
        summary = data_table.get("summary")
        if summary:
            summary_row = header_row + len(rows) + 2
            ws.cell(row=summary_row, column=1, value="總結").font = self.label_font
            ws.cell(row=summary_row, column=1).border = self.border
            ws.cell(row=summary_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
            summary_cell = ws.cell(row=summary_row, column=2, value=summary)
            summary_cell.font = self.content_font
            summary_cell.border = self.border
            summary_cell.alignment = Alignment(vertical="center", wrap_text=True)
            if headers:
                ws.merge_cells(start_row=summary_row, start_column=2, end_row=summary_row, end_column=len(headers))
            ws.row_dimensions[summary_row].height = 40

        # 自動調整欄寬
        for col in range(1, len(headers) + 1):
            col_letter = get_column_letter(col)
            max_length = 0
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    cell_length = len(str(cell.value)) * 1.2
                    if cell_length > max_length:
                        max_length = cell_length
            adjusted_width = max(12, min(max_length + 4, 60))
            ws.column_dimensions[col_letter].width = adjusted_width

        # 設置列高
        for row in range(1, ws.max_row + 1):
            ws.row_dimensions[row].height = 20

        # 生成檔名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"數據表格_{timestamp}_{uuid.uuid4().hex[:6]}.xlsx"
        filepath = os.path.join(self.output_dir, filename)

        wb.save(filepath)
        return filename

excel_generator = ExcelGenerator()